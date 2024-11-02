#!/usr/bin/env python
# coding: utf-8

# In[2]:

# In the program there are improvements:
# - Hidden number using getpass
# - Filters for charts by players who have played
# - Min_max module
# - Change of guessing range, so you can switch the range from 1-1000 to any maximum number

# Importing modules
import random 
import openpyxl 
import matplotlib.pyplot as plt 
import getpass 
import min_max
import os

# Define the path to the Excel file using a relative path
excel_path = os.path.join(os.path.dirname(__file__), 'records.xlsx')
# Load the workbook
excel1 = openpyxl.load_workbook(excel_path)

#-----------------------------------------------------------------

# Initial values
guess = 0
number = 0 
Stats = excel1['Stats'] 
difficulty = 0 
game_mode = 0 
Ys = [] 
Xs = [] 
stat = 0 
diff_num = 1000 #
#-----------------------------------------------------------------

# Dictionaries
game_modes = {"1": "1. Single Player Game", "2": "2. Two Player Game", "3": "3. Statistics", "4": "4. Exit"} 
difficulties = {"1": "1. Easy (20 attempts)", "2": "2. Medium (12 attempts)", "3": "3. Hard (5 attempts)"} 
stat_menu = {"1": "1 - View number of winners and losers", "2": "2 - Filter the number of times a player has played"} 
diff_range_menu = {"1": "1 - Default range from 1 - 1000", "2": "2 - Change the maximum range"} 
levels = {1: 20, 2: 12, 3: 5}
#-----------------------------------------------------------------

# Functions
def menu(): 
    print("--------------------------------------") 
    print(game_modes["1"]) 
    print(game_modes["2"]) 
    print(game_modes["3"]) 
    print(game_modes["4"]) 
    print("--------------------------------------")

def submenu(): 
    print("Select the difficulty level") 
    print(difficulties["1"]) 
    print(difficulties["2"]) 
    print(difficulties["3"]) 
    print("--------------------------------------")

def stat_menu_display(): 
    print("Choose the type of statistic to view") 
    print(stat_menu["1"]) 
    print(stat_menu["2"]) 
    print("--------------------------------------")

def difficulty_range_menu(): 
    print("The default game is to guess a number between 1 - 1000, but you can change the range:") 
    print(diff_range_menu["1"]) 
    print(diff_range_menu["2"]) 
    print("--------------------------------------")

def single_player(): 
    number = random.randrange(1, diff_num) 
    return number

def multiplayer(): 
    number = 0 
    while number < 1 or number > diff_num: 
        print("Player 1, please enter a number between 1 and " + str(int(diff_num))) 
        number = int(getpass.getpass("Player 1, enter the number to guess: ")) 
        print("--------------------------------------")
    return number

def game(): 
    if game_mode == 1: 
        number = single_player() 
    elif game_mode == 2: 
        number = multiplayer() 
        
    print("Now the guesser needs to guess the number") 
    for i in range(levels[difficulty]): 
        guess = int(input()) 
        if number > guess: 
            print("--------------------------------------") 
            print("The number to guess is higher than your guess") 
        elif number < guess: 
            print("--------------------------------------") 
            print("The number to guess is lower than your guess")  
        else: 
            print("--------------------------------------")
            print("Congratulations! You guessed the number!")
            print("Enter your name as the guessing champion")
            name = input()
            record = int(i + 1) 
            Stats.append([record] + [name] + ["winner"])
            excel1.save(r'records.xlsx')
            print("The game is over")
            print("--------------------------------------") 
            return 
        
    print("Game Over! You ran out of attempts!") 
    print("Enter your name") 
    name = input() 
    record = int(i + 1) 
    Stats.append([record] + [name] + ["loser"])
    excel1.save(r'records.xlsx')
    print("The game is over") 
    print("--------------------------------------") 
    return

def statistics():
    print("Statistics") 
    print("--------------------------------------")
    stat_menu_display()  
    stat = min_max.minimo_maximo(1, 2)
    if stat == 1: 
        print('Selected: ' + stat_menu[str(stat)])
        winners = 0 
        losers = 0 
        for row in Stats.iter_rows(max_row=Stats.max_row): 
            status_cell = row[2] 
            status = status_cell.value

            if status == "winner":
                winners += 1
            elif status == "loser":
                losers += 1

        x = ['Winners', 'Losers']
        y = [winners, losers]

        plt.bar(x, y)
        plt.show()

    elif stat == 2:
        print("Enter the name of the player you are searching for")
        print('Selected: ' + stat_menu[str(stat)])
        player_name = input()
        player_count = 0
        for row in Stats.iter_rows(max_row=Stats.max_row):
            status_cell = row[1]
            status = status_cell.value
    
            if status == player_name:
                player_count += 1

        x = [player_name]
        y = [player_count]
    
        plt.bar(x, y)
        plt.show()

def exit_game(): 
    print("Exiting the game")
#-----------------------------------------------------------------

# Main Program
while game_mode != 4: 
    difficulty = 0 
    print("Welcome to the Number Guessing Game! Choose either single-player or two-player mode.")
    menu()
    game_mode = min_max.minimo_maximo(1, 4)
    print('Selected: ' + game_modes[str(game_mode)])

    if game_mode <= 2:
        submenu()
        difficulty = min_max.minimo_maximo(1, 3)
        print('Selected: ' + difficulties[str(difficulty)])
    if game_mode <= 2:    
        range_choice = 0
        while 1 > range_choice or range_choice > 2:
            difficulty_range_menu()
            print("Please select either option 1 or 2")
            range_choice = int(input())
        if range_choice == 2:
            diff_num = 0
            while diff_num < 1:
                print("Enter a number greater than 1 for the guessing range")
                diff_num = int(input())
                print(diff_num)
        
    if game_mode == 1:
        game()
    elif game_mode == 2:
        game()
    elif game_mode == 3:
        statistics()
exit_game()